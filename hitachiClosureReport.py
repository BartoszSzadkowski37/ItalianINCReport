import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime
import Incident
import pyinputplus as pyip

#function to get letters of specific columns
def getLetters():
    global snowNumberCol
    global shortDescCol
    global descCol
    global resolvedCol
    global resolutionInfoCol
    global locationCol
    for i in range(1, sheet.max_column + 1):
        columnValue = sheet[get_column_letter(i) + '1'].value
        columnLetter = get_column_letter(i) 
        if columnValue=='Number':
           snowNumberCol = columnLetter
        elif columnValue=='Short Description':
            shortDescCol = columnLetter
        elif columnValue=='Description':
            descCol = columnLetter
        elif columnValue=='Resolved':
            resolvedCol = columnLetter
        elif columnValue=='Resolution Information':
            resolutionInfoCol = columnLetter
        elif columnValue=='Location':
            locationCol = columnLetter

            
    return

# This function will filter data by date and create array of INC elements
def dataFilter(lastReportDate):
    global incidents
    # loop through rows
    for row in range(2, sheet.max_row + 1):
        currentRowResolvedDate = sheet[resolvedCol+str(row)].value
        currentLocation = sheet[locationCol+str(row)].value
        if currentRowResolvedDate > lastReportDate and  currentLocation == 'HRAIL.IT.Napoli.Hitachi Rail Spa':
            currentSnowNumber = sheet[snowNumberCol+str(row)].value
            currentShortDesc = sheet[shortDescCol+str(row)].value
            currentDesc = sheet[descCol+str(row)].value
            currentResultInfo = sheet[resolutionInfoCol+str(row)].value
            currentIncident = Incident.Incident(currentSnowNumber, currentShortDesc, currentDesc, currentRowResolvedDate, currentResultInfo, currentLocation)


            currentIncident.getHRIandREF()

            if currentIncident.REF == '' or currentIncident.hriNumber == '':
                del currentIncident
                
                #IF INCIDENT DOES NOT HAVE REF OR HRI NUMBER IT SEEMS THAT IS TICKET NOT FROM HRI AND 
                #SHOULD NOT BE INCLUDE IN THE REPORT IN THAT CASE DELETE THIS INC OBJECT
            else:
            #ELSE ADD THIS INFO TO THE TEMPLATE
                currentIncident.createTheTemplate()
                incidents.append(currentIncident)
                #f = open('incidents.txt', 'a')
                #print(currentIncident.resolInfo)
                #print(type(currentIncident.resolInfo))
                #f.write('\n\n----------'+str(currentIncident.snowNumber)+'----------\n\n')
                #f.write('---------- Resolution Template ----------\n')
                #f.write(currentIncident.resolutionTemplate + '\n\n')
                #f.write('---------- Resolution Info ----------\n')
                #f.write(currentIncident.resolInfo + '\n')
                #f.write('***********************************\n\n\n')
                #f.write(str(currentIncident.snowNumber) + currentIncident.resolutionTemplate + '\n' + '-------RESOLUTION INFO------\n' + currentIncident.resolInfo + '\n')
                #f.close()
        
#function to export filtered date to the excel file
def exportToExcel(incidents):
    finalWb = openpyxl.Workbook()
    finalWb.sheetnames
    finalsheet = finalWb.active

    finalsheet['A1'] = 'SNOW Number'
    finalsheet['B1'] = 'HRI Number'
    finalsheet['C1'] = 'Resolution Info'
    finalsheet['D1'] = 'Resolution Template'

    for row in range(len(incidents)):
        finalsheet['A' + str(row+2)] = incidents[row].snowNumber
        finalsheet['B' + str(row+2)] = incidents[row].hriNumber
        finalsheet['C' + str(row+2)] = incidents[row].resolInfo
        finalsheet['D' + str(row+2)] = incidents[row].resolutionTemplate

    finalWb.save('ItalianIncidentsReport.xlsx')



#opening the workbook
wb=openpyxl.load_workbook('incident.xlsx')
sheet=wb['Page 1']

#creating variables for columns' letters
snowNumberCol=''
shortDescCol=''
descCol=''
resolvedCol=''
resolutionInfoCol=''
locationCol=''

#getting columns' letters 
getLetters()

#array of incident objects
incidents = []


lastReportDate = pyip.inputDatetime('Date of last report: ', formats=['%Y/%m/%d %H:%M'])

#filter incidents by date provided by user and appending incidents array by filtered objects
dataFilter(lastReportDate)


exportToExcel(incidents)